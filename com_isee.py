# -*- coding: utf-8 -*-
"""
================================================================================
SCRIPT : scripts/com_isee/com_isee.py
================================================================================
Génère et envoie automatiquement le relevé de prix HT mensuel pour l'ISEE.

Étapes :
  1. Charge article.dbf via SMB
  2. Filtre les références suivies par l'ISEE
  3. Génère un fichier Excel formaté
  4. Envoie le fichier par email aux destinataires

Auteur  : Stoyann - support QC
Date    : 2026-05-18
================================================================================
"""

import sys
import locale
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

# =====================================================
# CHEMINS & IMPORT DES MODULES
# =====================================================
# Racine dev/ = 3 niveaux au-dessus de ce script
ROOT        = Path(__file__).resolve().parent.parent.parent
MODULES_DIR = ROOT / "modules"

sys.path.insert(0, str(MODULES_DIR))
from _loader import load_module

# Chargement des modules
_dbf    = load_module("dbf-loader",          "dbf_loader.py")
_fm     = load_module("file-manager",        "file_manager.py")
_mailer = load_module("quincaillerie-mailer", "mailer.py")
_log    = load_module("logger-manager",      "logger.py")

# Fonctions exposées
get_dbf                = _dbf.get_dbf
generer_nom_fichier    = _fm.generer_nom_fichier
generer_chemin_fichier = _fm.generer_chemin_fichier
envoyer_email          = _mailer.envoyer_email
init_logger            = _log.init_logger

# =====================================================
# INITIALISATION LOGGER
# =====================================================
logger = init_logger("com_isee")

# =====================================================
# CONFIGURATION
# =====================================================
SOC = "QC"

MAIL_MAINTENANCE = [
    "support@quincaillerie.nc"
]

MAIL_TARGET = [
    "support@quincaillerie.nc",
    "exploitation@quincaillerie.nc",
    "n.leroux@quincaillerie.nc",
    "indices@isee.nc"
]

# Références articles suivies par l'ISEE
NARTS = [
    "710092", "760043", "760041", "761467", "760200", "760049",
    "760403", "760414", "210255", "120139", "590219", "833848",
    "820097", "870145", "870127", "760251", "760105", "761336",
    "240153", "790449", "160014", "160890"
]

# =====================================================
# FONCTIONS
# =====================================================
def _set_locale_fr():
    """Tente d'activer la locale française (Windows ou Linux)."""
    for loc in ("fr_FR.UTF-8", "French_France.1252", "fr_FR", "French"):
        try:
            locale.setlocale(locale.LC_TIME, loc)
            return
        except locale.Error:
            continue
    logger.warning("Locale française non disponible, utilisation de la locale système")


def generate_excel(df: pd.DataFrame, report_date: datetime) -> tuple:
    """
    Génère le fichier Excel formaté.

    Retourne (file_path, file_name)
    """
    _set_locale_fr()

    date_file = report_date.strftime("%b-%y").lower()
    file_name = f"{SOC}_comISEE_{date_file}.xlsx"
    file_path = generer_chemin_fichier(file_name)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ISEE")
        ws = writer.sheets["ISEE"]

        # Style en-tête
        fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font  = Font(bold=True, color="FFFFFF", size=11)
        align = Alignment(horizontal="center")

        for col_idx in range(1, len(df.columns) + 1):
            cell            = ws.cell(row=1, column=col_idx)
            cell.fill       = fill
            cell.font       = font
            cell.alignment  = align

        # Largeur automatique
        for col in ws.columns:
            max_len    = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 4

    logger.info(f"📁 Fichier généré : {file_path}")
    return file_path, file_name


def send_mails(file_path: Path, report_date: datetime) -> None:
    """Envoie le fichier Excel par email à tous les destinataires."""
    _set_locale_fr()

    date_mail = report_date.strftime("%B %Y").upper()
    sujet     = f"[{SOC}] - Relevé ISEE prix HT - {date_mail}"

    corps_html = f"""
    <html>
    <head>
        <style>
            body    {{ font-family: Arial, sans-serif; color: #333; }}
            .header {{ background-color: #4472C4; color: white; padding: 10px;
                       text-align: center; font-size: 18px; font-weight: bold; }}
            .content {{ margin: 20px; font-size: 14px; }}
            .footer  {{ margin: 20px; font-size: 12px; color: #777; }}
        </style>
    </head>
    <body>
        <div class="header">Relevé ISEE prix HT — {date_mail}</div>
        <div class="content">
            <p>Bonjour,</p>
            <p>Veuillez trouver en pièce jointe le relevé des prix HT
               pour <strong>{date_mail}</strong>.</p>
            <p>Cordialement,<br>Service Analyse Commercial</p>
        </div>
        <div class="footer">
            <p>Ce message est envoyé automatiquement, merci de ne pas y répondre.</p>
        </div>
    </body>
    </html>
    """

    # Fusion et dédoublonnage des destinataires
    tous = list(set(MAIL_MAINTENANCE + MAIL_TARGET))

    for dest in tous:
        logger.info(f"📤 Envoi à {dest}...")
        ok = envoyer_email(
            destinataire=dest,
            sujet=sujet,
            corps=corps_html,
            html=True,
            chemin_piece_jointe=str(file_path)
        )
        if ok:
            logger.info(f"✅ Mail envoyé → {dest}")
        else:
            logger.error(f"❌ Échec envoi → {dest}")


# =====================================================
# MAIN
# =====================================================
def main():
    logger.info("=" * 60)
    logger.info("DÉMARRAGE — com_isee")
    logger.info("=" * 60)

    # Mois du rapport = mois précédent
    today       = datetime.now()
    report_date = (today.replace(day=1) - timedelta(days=1))
    logger.info(f"Rapport pour : {report_date.strftime('%B %Y')}")

    try:
        # 1. Chargement DBF
        logger.info("Chargement article.dbf...")
        df = get_dbf("qc/article.dbf")
        logger.info(f"{len(df)} lignes chargées")

        # 2. Filtrage
        df_filtered = df[df["NART"].astype(str).str.strip().isin(NARTS)]
        logger.info(f"{len(df_filtered)} articles ISEE trouvés")

        if df_filtered.empty:
            logger.warning("Aucun article trouvé — vérifiez les références NARTS")

        # 3. Sélection et renommage des colonnes
        df_result = df_filtered[["NART", "DESIGN", "PVTE"]].rename(columns={
            "NART":   "Réf_Mag",
            "DESIGN": "Produit",
            "PVTE":   "PVTE_HT"
        })

        # 4. Génération Excel
        file_path, _ = generate_excel(df_result, report_date)

        # 5. Envoi emails
        send_mails(file_path, report_date)

        logger.info("=" * 60)
        logger.info("✅ Script terminé avec succès")
        logger.info("=" * 60)

    except Exception as e:
        logger.error(f"❌ ERREUR CRITIQUE : {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
